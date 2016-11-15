import argparse
import json
import random
import re
import requests
from bs4 import BeautifulSoup
from lxml import etree
from openpyxl import Workbook


def obtain_courses_links(count=20):
    url = 'https://www.coursera.org/sitemap~www~courses.xml'
    xml = requests.get(url)
    root = etree.fromstring(xml.content)
    links = [link.text for link in root.iter('{*}loc') if 'learn' in link.text]
    random.shuffle(links)
    return links[:count]


def fetch_course_page(url):
    html = requests.get(url).content
    return BeautifulSoup(html, 'lxml')


def get_course_title(soup):
    title = soup.find('div', {'class': 'title'})
    return title.text if title else None


def get_course_lang(soup):
    table = soup.find('table', {'class': 'basic-info-table'})
    trs = table.findAll('tr')
    for tr in trs:
        if re.search(r'Lang', tr.text):
            key, value = [td.text.split()[0] for td in tr.findAll('td')]
            return value.rstrip(',')


def get_course_start_date(soup):
    script = soup.find('script', {'type': 'application/ld+json'})
    if script and re.search(r'startDate', script.text):
        return json.loads(script.text)['hasCourseInstance'][0]['startDate']


def count_course_duration(soup):
    return len(soup.findAll('div', {'class': 'week'}))


def get_course_rate(soup):
    rate = soup.find('div', {'class': 'ratings-text'})
    return rate.text if rate else None


def collect_course_info(url):
    soup = fetch_course_page(url)
    course = {
        'Title': get_course_title(soup),
        'Language': get_course_lang(soup),
        'Start Date': get_course_start_date(soup),
        'Weeks amount': count_course_duration(soup),
        'Course Rate': get_course_rate(soup),
        'URL': url
    }
    return course


def output_courses_info_to_xlsx(filepath, courses):
    headers = ['Title', 'Language', 'Start Date',
               'Weeks amount', 'Course Rate', 'URL']
    wb = Workbook()
    ws = wb.active
    ws.title = 'coursera'
    for j, col in enumerate(headers, 1):
        ws.cell(row=1, column=j).value = col
    for num, course in enumerate(courses, 2):
        for i, key in enumerate(headers, 1):
            if key == 'URL':
                ws.cell(row=num, column=i).value = 'more details →'
                ws.cell(row=num, column=i).hyperlink = course[key]
            else:
                ws.cell(row=num, column=i).value = course[key]
    wb.save(filename=filepath)
    return True


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Скрипт собирает информацию о разных курсах на Курсере')
    parser.add_argument('filepath', help='укажите файл в формате xlsx')
    args = parser.parse_args()
    links = obtain_courses_links(20)
    courses = [collect_course_info(link) for link in links]
    if output_courses_info_to_xlsx(args.filepath, courses):
        print('Data has been written to file: {}'.format(args.filepath))
    else:
        print('Something goes wrong :(')
